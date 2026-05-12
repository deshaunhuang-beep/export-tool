import csv
import pymongo
import pymongo.errors
from datetime import datetime, timedelta
import sys
import os
import traceback
import json

try:
    import openpyxl
except ImportError:
    pass

VERSION = "7.1.0-FullFeatures"

def get_base_path():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    else:
        return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = get_base_path()
CONFIG_FILE = os.path.join(BASE_DIR, "config.json")

def safe_date_format(dt_obj):
    if not dt_obj or not isinstance(dt_obj, datetime):
        return ""
    try:
        return (dt_obj + timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
    except:
        return ""

def load_or_ask_config():
    config = {"mongo_uri": "", "db_name": "", "app_id": ""}

    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                config.update(json.load(f))
            print(f"✅ 已加载本地配置 (路径: {CONFIG_FILE})")
        except:
            print("⚠️ 配置文件读取失败，将重新输入")

    print("\n--- 基础配置 ---")
    if not config["mongo_uri"]:
        while not config["mongo_uri"]:
            config["mongo_uri"] = input("请输入 MongoDB 连接地址: ").strip()
    else:
        mongo_in = input("MongoDB 地址 (回车保持现状): ").strip()
        if mongo_in: config["mongo_uri"] = mongo_in

    if not config["db_name"]:
        while not config["db_name"]:
            config["db_name"] = input("请输入数据库名称: ").strip()
    else:
        db_in = input(f"数据库名称 [{config['db_name']}] (回车保持现状): ").strip()
        if db_in: config["db_name"] = db_in

    if not config["app_id"]:
        while not config["app_id"]:
            app_in = input("请输入业务 AppID (必须为数字): ").strip()
            if app_in.isdigit(): config["app_id"] = int(app_in)
            else: print("❌ AppID 必须是数字")
    else:
        app_in = input(f"业务 AppID [{config['app_id']}] (回车保持现状): ").strip()
        if app_in:
            if app_in.isdigit(): config["app_id"] = int(app_in)
            else: print("❌ AppID 必须是数字，已保留原配置")

    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=4, ensure_ascii=False)

    return config

def get_int_input(prompt_text, default_val):
    val = input(prompt_text).strip()
    if not val:
        return default_val
    try:
        return int(val)
    except ValueError:
        print(f"   ⚠️ 输入无效，自动使用默认值: {default_val}")
        return default_val

def run_report_1_chongti(db, config, start_utc, end_utc, date_str):
    output_file = os.path.join(BASE_DIR, f"充提数据_{config['db_name']}_{date_str}.csv")
    print(f"\n[1/4] 正在拉取基础订单 ({date_str})...")
    
    pipeline = [
        {"$match": {"appID": config['app_id'], "updatedAt": {"$gte": start_utc, "$lt": end_utc}, 
                    "type": {"$in": ['pay', 'withdrawal']}, "status": {"$in": ['Completed', 'MockCompleted']}, 
                    "ignoreAnalysis": {"$ne": True}}},
        {"$group": {"_id": {"user": "$user", "channel": {"$ifNull": ["$channel", 0]}},
                    "firstStatus": {"$first": "$status"},
                    "存款次数": {"$sum": {"$cond": [{"$eq": ["$type", "pay"]}, 1, 0]}},
                    "存款金额": {"$sum": {"$cond": [{"$eq": ["$type", "pay"]}, {"$ifNull": ["$totalPrice", 0]}, 0]}},
                    "提款金额": {"$sum": {"$cond": [{"$eq": ["$type", "withdrawal"]}, {"$ifNull": ["$totalPrice", 0]}, 0]}}}}
    ]
    order_results = list(db["orders"].aggregate(pipeline, allowDiskUse=True))
    
    if not order_results:
        print("⚠️ 该时间段未找到数据。")
        return

    uids = [res['_id']['user'] for res in order_results]
    print(f"[2/4] 同步 {len(uids)} 个用户...")
    user_map = {u['_id']: u for u in db["users"].find({"_id": {"$in": uids}}, {"uid": 1, "meta.adChannel": 1, "createdAt": 1})}
    
    daily_pipeline = [
        {"$match": {"user": {"$in": uids}, "startAt": {"$gte": start_utc, "$lt": end_utc}}},
        {"$group": {"_id": "$user", "rewardCash": {"$sum": "$rewardCash"}, "betAmount": {"$sum": "$betAmount"}}}
    ]
    daily_map = {res['_id']: res for res in db["transactiondailies"].aggregate(daily_pipeline)}

    print(f"[3/4] 写入文件...")
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(["uid", "用户渠道", "注册日期", "是否模拟回调", "区间内存款金额", "区间内提款金额", "区间内存款次数", "区间内获得真金奖励", "区间内投注金额"])
        rows_to_write = []
        for doc in order_results:
            u_id = doc['_id']['user']
            u = user_map.get(u_id, {})
            d = daily_map.get(u_id, {})
            rows_to_write.append([
                u.get('uid', ''), "广告" if u.get('meta', {}).get('adChannel') else "自然裂变",
                safe_date_format(u.get('createdAt')),
                "是" if doc.get('firstStatus') == 'MockCompleted' else "否",
                doc.get('存款金额', 0), doc.get('提款金额', 0), doc.get('存款次数', 0),
                d.get('rewardCash', 0), d.get('betAmount', 0)
            ])
        writer.writerows(rows_to_write)
    print(f"✅ 完成: {os.path.abspath(output_file)}")

def run_report_2_shoucun(db, config, start_utc, end_utc, date_str):
    output_file = os.path.join(BASE_DIR, f"首存订单_{config['db_name']}_{date_str}.csv")
    print(f"\n[1/4] 正在统计该时间段所有充值用户 ({date_str})...")
    
    pay_pipeline = [
        {"$match": {"appID": config['app_id'], "updatedAt": {"$gte": start_utc, "$lt": end_utc}, 
                    "type": "pay", "status": "Completed", "ignoreAnalysis": {"$ne": True}}},
        {"$sort": {"updatedAt": 1, "_id": 1}},
        {"$group": {"_id": "$user", "次数": {"$sum": 1}, "总额": {"$sum": "$totalPrice"}, "第一笔": {"$first": "$totalPrice"}}}
    ]
    pay_results = list(db["orders"].aggregate(pay_pipeline, allowDiskUse=True))
    uids = [res['_id'] for res in pay_results]
    
    print(f"[2/4] 正在筛选在这几天内产生首次充值的用户...")
    shoucun_users = {u['_id']: u for u in db["users"].find({
        "_id": {"$in": uids},
        "meta.firstRechargeAt": {"$gte": start_utc, "$lt": end_utc}
    }, {"uid": 1, "meta.adChannel": 1, "meta.firstRechargeAt": 1, "createdAt": 1})}
    
    sc_uids = list(shoucun_users.keys())
    if not sc_uids:
        print("⚠️ 该时间段内无首存用户。")
        return

    print(f"[3/4] 抓取提款与日报数据 (共 {len(sc_uids)} 人)...")
    wd_pipeline = [{"$match": {"user": {"$in": sc_uids}, "type": "withdrawal", "status": "Completed", 
                               "updatedAt": {"$gte": start_utc, "$lt": end_utc}}},
                   {"$group": {"_id": "$user", "total": {"$sum": "$totalPrice"}}}]
    wd_map = {res['_id']: res['total'] for res in db["orders"].aggregate(wd_pipeline)}
    
    daily_pipeline = [
        {"$match": {"user": {"$in": sc_uids}, "startAt": {"$gte": start_utc, "$lt": end_utc}}},
        {"$group": {"_id": "$user", "rewardCash": {"$sum": "$rewardCash"}, "betAmount": {"$sum": "$betAmount"}}}
    ]
    daily_map = {res['_id']: res for res in db["transactiondailies"].aggregate(daily_pipeline)}

    print(f"[4/4] 导出报表...")
    headers = ["用户渠道", "uid", "注册日期", "首次充值日期", "区间内存款次数", "第一笔充值金额", "区间内总充值金额", "区间内提款金额", "区间内获得真金奖励", "区间内投注金额"]
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        rows_to_write = []
        for res in pay_results:
            uid = res['_id']
            if uid not in shoucun_users: continue
            u = shoucun_users[uid]
            d = daily_map.get(uid, {})
            rows_to_write.append([
                "广告" if u.get('meta', {}).get('adChannel') else "自然裂变",
                u.get('uid', ''), safe_date_format(u.get('createdAt')), safe_date_format(u.get('meta', {}).get('firstRechargeAt')),
                res.get('次数', 0), res.get('第一笔', 0), res.get('总额', 0), wd_map.get(uid, 0), d.get('rewardCash', 0), d.get('betAmount', 0)
            ])
        writer.writerows(rows_to_write)
    print(f"✅ 完成: {os.path.abspath(output_file)}")

def run_report_3_sms_recall(db, config, start_utc, end_utc, date_str):
    print(f"\n--- [书生计算 SMS 召回情况 ({date_str})] ---")
    file_name = input("请输入文件名 (支持 .csv 或 .xlsx，例如 target_users.xlsx): ").strip()
    
    if not os.path.isabs(file_name):
        file_name = os.path.join(BASE_DIR, file_name)
        
    if not os.path.exists(file_name):
        print(f"❌ 找不到文件: {file_name}，请确保它和本程序在同一个文件夹！")
        return

    target_uids = []
    ext = os.path.splitext(file_name)[1].lower()
    try:
        if ext == '.csv':
            with open(file_name, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                uid_key = next((k for k in reader.fieldnames if str(k).strip().lower() == 'uid'), None)
                if not uid_key: return print("❌ CSV 文件的表头里没有找到 'uid' 或 'UID'！请检查。")
                for row in reader:
                    if str(row[uid_key]).strip().isdigit(): target_uids.append(int(str(row[uid_key]).strip()))
        elif ext == '.xlsx':
            if 'openpyxl' not in sys.modules: return print("❌ 缺少 openpyxl 库。")
            wb = openpyxl.load_workbook(file_name, data_only=True)
            sheet = wb.active
            uid_col_idx = next((i for i, cell in enumerate(sheet[1], start=1) if cell.value and str(cell.value).strip().lower() == 'uid'), None)
            if not uid_col_idx: return print("❌ Excel 文件的【第一行】表头里没有找到 'uid' 或 'UID' 列！请检查。")
            for row_idx in range(2, sheet.max_row + 1):
                val = sheet.cell(row=row_idx, column=uid_col_idx).value
                if val is not None and str(val).strip().isdigit(): target_uids.append(int(str(val).strip()))
        else: return print(f"❌ 不支持的文件格式: {ext}")
    except Exception as e:
        return print(f"❌ 读取文件失败: {e}")

    target_uids = list(set(target_uids))
    if not target_uids: return print("⚠️ 文件里没有解析到任何有效的 UID。")
        
    print(f"✅ 成功从 {ext.upper()} 载入 {len(target_uids)} 个有效 UID。正在分析...")

    pipeline = [
        {"$match": {"uid": {"$in": target_uids}}},
        {"$group": {
            "_id": None, "totalUsers": {"$sum": 1},
            "activeUsers": {"$sum": {"$cond": [{"$and": [{"$gte": ["$updatedAt", start_utc]}, {"$lt": ["$updatedAt", end_utc]}]}, 1, 0]}},
            "rechargeUsers": {"$sum": {"$cond": [{"$and": [{"$ne": ["$meta.lastRechargeAt", None]}, {"$gte": ["$meta.lastRechargeAt", start_utc]}, {"$lt": ["$meta.lastRechargeAt", end_utc]}]}, 1, 0]}}
        }},
        {"$project": {"_id": 0, "totalUsers": 1, "activeUsers": 1, "rechargeUsers": 1, "rechargeRate": {"$cond": [{"$eq": ["$activeUsers", 0]}, 0, {"$divide": ["$rechargeUsers", "$activeUsers"]}]}}}
    ]
    result = list(db["users"].aggregate(pipeline, allowDiskUse=True))
    
    print("\n" + "🌟"*20 + f"\n  SMS 召回数据统计 ({date_str})\n" + "🌟"*20)
    if result:
        d = result[0]
        print(f"🔹 目标用户: {d.get('totalUsers',0):,} 人 | 活跃: {d.get('activeUsers',0):,} 人 | 充值: {d.get('rechargeUsers',0):,} 人 | 转化率: {d.get('rechargeRate',0):.2%}")
    else: print("⚠️ 没有查询到结果")
    print("🌟"*20)

def run_report_4_unrecharged_users(db, config, end_utc, end_date_str):
    output_file = os.path.join(BASE_DIR, f"注册未充值用户_{config['db_name']}_{end_date_str}.csv")
    print(f"\n[1/2] 正在筛选注册未充值用户 (截止到 {end_date_str} 23:59:59)...")

    query = {"role": {"$ne": "gm"}, "rechargeCount": 0, "updatedAt": {"$lt": end_utc}}
    projection = {"_id": 0, "uid": 1, "phone": 1, "email": 1, "updatedAt": 1}
    cursor = db["users"].find(query, projection, batch_size=5000)

    print(f"[2/2] 正在生成导出文件...")
    count, batch_data, batch_size = 0, [], 10000
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(["uid", "手机号", "邮箱", "最后活跃时间(东八区)"])
        for doc in cursor:
            raw_phone = doc.get('phone', '') or ''
            batch_data.append([doc.get('uid', ''), f"\t{raw_phone}" if raw_phone else "", doc.get('email', '') or '', safe_date_format(doc.get('updatedAt'))])
            count += 1
            if count % batch_size == 0:
                writer.writerows(batch_data)
                batch_data = []
                print(f"  已处理 {count} 条...")
        if batch_data: writer.writerows(batch_data)
    print(f"✅ 导出成功！共计 {count} 名注册未充值用户。\n文件位置: {os.path.abspath(output_file)}")

def run_report_5_custom_users(db, config, end_utc, end_date_str):
    output_file = os.path.join(BASE_DIR, f"圈选用户群_{config['db_name']}_{end_date_str}.csv")
    
    print("\n" + "="*40)
    print(" 🎯 请输入筛选条件 (直接回车将使用默认值)")
    print("="*40)

    rc_min = get_int_input("▶ 1. 最小充值次数 (默认 1): ", 1)
    rc_max = get_int_input("▶    最大充值次数 (默认 1): ", 1)
    if rc_min > rc_max: rc_min, rc_max = rc_max, rc_min

    print("-" * 30)
    cash_min = get_int_input("▶ 2. 最小充值金额 (默认 100): ", 100)
    cash_max = get_int_input("▶    最大充值金额 (默认 999999): ", 999999)
    if cash_min > cash_max: cash_min, cash_max = cash_max, cash_min

    print("-" * 30)
    off_min = get_int_input("▶ 3. 最小离线天数 (默认 3): ", 3)
    off_max = get_int_input("▶    最大离线天数 (默认 9999): ", 9999)
    if off_min > off_max: off_min, off_max = off_max, off_min

    print("-" * 30)
    bal_min = get_int_input("▶ 4. 最小账户余额 (默认 0): ", 0)
    bal_max = get_int_input("▶    最大账户余额 (默认 5): ", 5)
    if bal_min > bal_max: bal_min, bal_max = bal_max, bal_min

    print("="*40)

    max_login_time = end_utc - timedelta(days=off_min)
    min_login_time = end_utc - timedelta(days=off_max)

    print(f"\n[1/3] 条件确认完毕:")
    print(f"      - 充值次数: {rc_min} 至 {rc_max} 次")
    print(f"      - 充值金额: {cash_min} 至 {cash_max} 元")
    print(f"      - 账户余额: {bal_min} 至 {bal_max} 元")
    print(f"      - 离线天数: {off_min} 至 {off_max} 天 (最后登录晚于 {min_login_time.strftime('%Y-%m-%d')} 且早于 {max_login_time.strftime('%Y-%m-%d')})")

    query = {
        "role": {"$ne": "gm"},
        "rechargeCount": {"$gte": rc_min, "$lte": rc_max},
        "rechargeCash": {"$gte": cash_min, "$lte": cash_max},
        "cash": {"$gte": bal_min, "$lte": bal_max},
        "latestLoginAt": {"$gte": min_login_time, "$lt": max_login_time}
    }
    
    projection = {"_id": 1, "uid": 1, "phone": 1, "email": 1, "rechargeCash": 1, "latestLoginAt": 1, "cash": 1}

    print(f"\n[2/3] 正在数据库中圈选符合条件的用户，请稍候...")
    cursor = db["users"].find(query, projection, batch_size=5000)

    print(f"[3/3] 正在匹配 KYC 数据并生成导出文件...")
    
    count = 0
    batch_size = 5000
    docs_cache = []
    user_ids_cache = []
    
    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow(["uid", "手机号", "最后登陆时间(东八区)", "邮箱", "充值总金额", "账户余额", "KYC手机号"])

        def process_and_write_batch(users, uids):
            wallets = db["wallets"].find({"user": {"$in": uids}}, {"user": 1, "banks": 1})
            kyc_map = {}
            for w in wallets:
                banks = w.get("banks", [])
                if banks and isinstance(banks, list) and len(banks) > 0:
                    kyc_map[w["user"]] = banks[0].get("phone", "")

            rows = []
            for u in users:
                raw_phone = u.get('phone', '') or ''
                phone = f"\t{raw_phone}" if raw_phone else ""
                raw_kyc = kyc_map.get(u['_id'], "")
                kyc_phone = f"\t{raw_kyc}" if raw_kyc else ""
                
                rows.append([
                    u.get('uid', ''), phone, safe_date_format(u.get('latestLoginAt')),
                    u.get('email', '') or '', u.get('rechargeCash', 0), u.get('cash', 0), kyc_phone
                ])
            writer.writerows(rows)

        for doc in cursor:
            docs_cache.append(doc)
            user_ids_cache.append(doc['_id'])
            count += 1
            if len(docs_cache) >= batch_size:
                process_and_write_batch(docs_cache, user_ids_cache)
                docs_cache.clear()
                user_ids_cache.clear()
                print(f"  已处理并导出 {count} 条精准用户数据...")

        if docs_cache:
            process_and_write_batch(docs_cache, user_ids_cache)

    print(f"✅ 导出成功！共计圈选出 {count} 名符合条件的用户。\n文件位置: {os.path.abspath(output_file)}")

def run_report_6_inactive_rechargers(db, config, end_utc, date_str):
    """逻辑 6: 导出距上次充值超过 X 天的客户数据 (找回并升级了KYC匹配)"""
    output_file = os.path.join(BASE_DIR, f"未复充沉睡用户_{config['db_name']}_{date_str}.csv")
    
    print("\n" + "="*40)
    days = get_int_input("▶ 请输入距离上次充值的天数 (默认 3): ", 3)
    print("="*40)

    target_date = end_utc - timedelta(days=days)
    print(f"\n[1/3] 正在筛选距上次充值超过 {days} 天的用户...")
    print(f"      (最后充值时间早于东八区: {(target_date + timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')})")

    query = {
        "role": {"$ne": "gm"},
        "rechargeCount": {"$gt": 0},
        "$or": [
            {"meta.lastRechargeAt": {"$lt": target_date}},
            {"meta.lastRechargeAt": {"$exists": False}},
            {"meta.lastRechargeAt": None}
        ]
    }

    projection = {
        "_id": 1, "uid": 1, "phone": 1, "email": 1, "rechargeCash": 1,
        "rechargeCount": 1, "withdrawCompletedCash": 1, "withdrawCash": 1,
        "withdrawCompletedCount": 1, "withdrawCount": 1, "latestLoginAt": 1,
        "meta.lastRecharge": 1, "meta.lastRechargeAt": 1
    }

    cursor = db["users"].find(query, projection, batch_size=5000)

    print(f"[2/3] 正在匹配 KYC 认证数据并生成导出文件...")

    count = 0
    batch_size = 5000
    docs_cache = []
    user_ids_cache = []

    with open(output_file, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f)
        writer.writerow([
            "uid", "手机号", "KYC手机号", "邮箱", "最后充值时间(东八区)", "距今天数",
            "充值总金额", "充值总次数", "最后充值金额", "提款总金额", "提款总次数", "最后登陆时间(东八区)"
        ])

        def process_and_write_batch(users, uids):
            wallets = db["wallets"].find({"user": {"$in": uids}}, {"user": 1, "banks": 1})
            kyc_map = {}
            for w in wallets:
                banks = w.get("banks", [])
                if banks and isinstance(banks, list) and len(banks) > 0:
                    kyc_map[w["user"]] = banks[0].get("phone", "")

            rows = []
            for u in users:
                raw_phone = u.get('phone', '') or ''
                phone = f"\t{raw_phone}" if raw_phone else ""
                
                raw_kyc = kyc_map.get(u['_id'], "")
                kyc_phone = f"\t{raw_kyc}" if raw_kyc else ""

                meta = u.get('meta', {})
                last_rech_at = meta.get('lastRechargeAt')

                days_diff = ""
                if last_rech_at and isinstance(last_rech_at, datetime):
                    days_diff = (end_utc - last_rech_at).days

                rows.append([
                    u.get('uid', ''), phone, kyc_phone, u.get('email', '') or '',
                    safe_date_format(last_rech_at), days_diff,
                    u.get('rechargeCash', 0), u.get('rechargeCount', 0), meta.get('lastRecharge', 0),
                    u.get('withdrawCompletedCash', u.get('withdrawCash', 0)),
                    u.get('withdrawCompletedCount', u.get('withdrawCount', 0)),
                    safe_date_format(u.get('latestLoginAt'))
                ])
            writer.writerows(rows)

        for doc in cursor:
            docs_cache.append(doc)
            user_ids_cache.append(doc['_id'])
            count += 1
            if len(docs_cache) >= batch_size:
                process_and_write_batch(docs_cache, user_ids_cache)
                docs_cache.clear()
                user_ids_cache.clear()
                print(f"  已处理并导出 {count} 条精准用户数据...")

        if docs_cache:
            process_and_write_batch(docs_cache, user_ids_cache)

    print(f"[3/3] ✅ 导出成功！共计圈选出 {count} 名符合条件的用户。\n文件位置: {os.path.abspath(output_file)}")


def run_report_7_phone_payment_behavior(db, config, start_utc, end_utc, date_str):
    print(f"\n--- [根据手机号码查询用户付费行为 ({date_str})] ---")
    
    txt_files = [f for f in os.listdir(BASE_DIR) if f.lower().endswith('.txt')]
    
    if not txt_files:
        print("\n❌ 当前目录下没有找到任何 .txt 文件！")
        print("   请先将包含手机号的 .txt 文件放入本程序所在的文件夹。")
        return
        
    print("\n📂 找到以下 .txt 文件，请选择：")
    for idx, f in enumerate(txt_files):
        print(f"  [{idx + 1}] {f}")
        
    file_idx = get_int_input("\n▶ 请输入文件编号进行选择: ", 1)
    if file_idx < 1 or file_idx > len(txt_files):
        print("❌ 无效的选择，操作取消。")
        return
        
    selected_file = os.path.join(BASE_DIR, txt_files[file_idx - 1])
    print(f"\n📄 已选择文件: {txt_files[file_idx - 1]}")
    
    phone_set = set()
    try:
        with open(selected_file, 'r', encoding='utf-8-sig') as f:
            for line in f:
                p = line.strip()
                if p: phone_set.add(p)
    except Exception as e:
        print(f"❌ 读取文件失败: {e}")
        return
        
    total_phones = len(phone_set)
    if total_phones == 0:
        print("⚠️ 文件中没有找到任何有效的手机号。")
        return
        
    print(f"✅ 成功读取了 {total_phones} 个去重后的手机号。正在匹配系统用户...")
    
    search_phones = []
    for p in phone_set:
        search_phones.append(p)
        if not p.startswith('+'):
            search_phones.append(f"+63-{p}")
            search_phones.append(f"+63{p}")
            if p.startswith('0'):
                search_phones.append(f"+63-{p[1:]}")
                
    users_cursor = db["users"].find({"phone": {"$in": search_phones}}, {"_id": 1})
    matched_user_ids = [u['_id'] for u in users_cursor]
        
    matched_count = len(matched_user_ids)
    if matched_count == 0:
        print("\n⚠️ 匹配失败：在数据库中未能匹配到对应的系统用户。")
        return
        
    print(f"✅ 成功匹配到 {matched_count} 个系统用户 (UID)。正在查询付费行为...")
    
    pay_pipeline = [
        {"$match": {
            "appID": config['app_id'], 
            "user": {"$in": matched_user_ids},
            "type": "pay", 
            "status": {"$in": ['Completed', 'MockCompleted']},
            "ignoreAnalysis": {"$ne": True},
            "updatedAt": {"$gte": start_utc, "$lt": end_utc}
        }},
        {"$group": {
            "_id": "$user",
            "totalAmount": {"$sum": "$totalPrice"}
        }}
    ]
    
    pay_results = list(db["orders"].aggregate(pay_pipeline, allowDiskUse=True))
    
    paying_users_count = len(pay_results)
    total_paid_amount = sum(res.get('totalAmount', 0) for res in pay_results)
    
    print("\n" + "🌟"*25)
    print(f"  📊 手机号渠道转化质量分析 ({date_str})")
    print("🌟"*25)
    print(f"🔹 1. 文件提取手机号总数 : {total_phones:,} 个")
    print(f"🔹 2. 成功定位系统用户数 : {matched_count:,} 人")
    print(f"🔹 3. 期间产生付费用户数 : {paying_users_count:,} 人")
    
    if matched_count > 0:
        conversion_rate = paying_users_count / matched_count
        print(f"🔹 4. 付费转化率 (付费/匹配): {conversion_rate:.2%}")
    
    if paying_users_count > 0:
        arpu = total_paid_amount / paying_users_count
        print(f"🔹 5. 期间累积总付费金额   : {total_paid_amount:,.2f}")
        print(f"🔹 6. 期间付费用户 ARPU 值 : {arpu:,.2f}")
    else:
        print("🔹 期间无任何付费行为")
    print("🌟"*25)

def main():
    print("=" * 50)
    print(f"      运营数据自动化导出工具 v{VERSION}")
    print("=" * 50)
    print(f"运行目录: {BASE_DIR}")

    try:
        config = load_or_ask_config()

        if not config["mongo_uri"] or not config["db_name"] or not config["app_id"]:
            print("❌ 核心配置不完整，程序退出。")
            return

        print("\n--- 请选择要执行的功能 ---")
        print("[1] 导出 - 充提数据")
        print("[2] 导出 - 首存订单")
        print("[3] 打印 - 书生计算SMS召回情况 (支持 .csv / .xlsx)")
        print("[4] 导出 - 书生筛选注册未充值用户 (用于拉新/激活)")
        print("[5] 导出 - 查询指定条件用户群 (用于精准圈选)")
        print("[6] 导出 - 距上次充值超过X天的客户数据 (促活/召回)")
        print("[7] 打印 - 根据手机号码查询用户付费行为 (需本地txt)")
        
        choice = input(">> ").strip()

        print("\n--- 日期范围设置 ---")
        start_in = input("请输入起始日期 (YYYY-MM-DD, 回车默认昨天): ").strip()
        if not start_in:
            start_date = datetime.now() - timedelta(days=1)
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
        else:
            try: start_date = datetime.strptime(start_in, "%Y-%m-%d")
            except ValueError: return print("❌ 日期格式错误")

        end_in = input(f"请输入结束日期 (YYYY-MM-DD, 回车默认与起始日期相同 [{start_date.strftime('%Y-%m-%d')}]): ").strip()
        if not end_in: end_date = start_date
        else:
            try: end_date = datetime.strptime(end_in, "%Y-%m-%d")
            except ValueError: return print("❌ 日期格式错误")
        
        if end_date < start_date: return print("❌ 结束日期不能早于起始日期！")

        start_utc = start_date - timedelta(hours=8)
        end_utc = end_date + timedelta(days=1) - timedelta(hours=8)
        date_str = start_date.strftime("%Y-%m-%d") if start_date == end_date else f"{start_date.strftime('%Y-%m-%d')}_至_{end_date.strftime('%Y-%m-%d')}"
        
        if choice in ['4', '5', '6']:
            print(f"\n⏳ 统计截止时间基准 (北京时间): {end_date.strftime('%Y-%m-%d 23:59:59')}")
        else:
            print(f"\n⏳ 统计时间范围 (北京时间): {start_date.strftime('%Y-%m-%d 00:00:00')} -> {end_date.strftime('%Y-%m-%d 23:59:59')}")

        print("\n正在连接 MongoDB...")
        client = pymongo.MongoClient(config['mongo_uri'], serverSelectionTimeoutMS=5000)
        client.admin.command('ping')
        db = client[config['db_name']]
        print("✅ MongoDB 连接成功")

        if choice == '1': run_report_1_chongti(db, config, start_utc, end_utc, date_str)
        elif choice == '2': run_report_2_shoucun(db, config, start_utc, end_utc, date_str)
        elif choice == '3': run_report_3_sms_recall(db, config, start_utc, end_utc, date_str)
        elif choice == '4': run_report_4_unrecharged_users(db, config, end_utc, end_date.strftime('%Y-%m-%d'))
        elif choice == '5': run_report_5_custom_users(db, config, end_utc, end_date.strftime('%Y-%m-%d'))
        elif choice == '6': run_report_6_inactive_rechargers(db, config, end_utc, end_date.strftime('%Y-%m-%d'))
        elif choice == '7': run_report_7_phone_payment_behavior(db, config, start_utc, end_utc, date_str)
        else: print("❌ 无效选择")

    except Exception as e:
        print("\n❌ 程序运行发生未知错误")
        traceback.print_exc()
    finally:
        print("\n" + "=" * 50)
        input("程序已结束，按 [回车键] 关闭窗口...")

if __name__ == "__main__":
    main()
